#!/usr/bin/env python

from __future__ import division, print_function

import re
from . import dna, gen9
from .. import scripting
from .digest import restriction_sites
from pathlib import Path

def translate(dna_seq, strip_stop=True):
    from more_itertools import chunked
    protein_seq = ''.join(
            dna.genetic_code[''.join(codon).upper()]
            for codon in chunked(dna_seq, 3)
    )
    if strip_stop:
        protein_seq = protein_seq.strip('.')
    return protein_seq

def reverse_translate(
        protein_seq,
        template_dna=None, leading_seq=None, trailing_seq=None,
        forbidden_seqs=(), include_stop=True, manufacturer=None):
    """
    Generate a well-behaved DNA sequence from the given protein sequence.  If a 
    template DNA sequence is specified, the returned DNA sequence will be as 
    similar to it  as possible.  Any given restriction sites will not be 
    present in the sequence.  And finally, the given leading and trailing 
    sequences will be appropriately concatenated.
    """

    if manufacturer == 'gen9':
        forbidden_seqs += gen9.reserved_restriction_sites

    leading_seq = restriction_sites.get(leading_seq, leading_seq or '')
    trailing_seq = restriction_sites.get(trailing_seq, trailing_seq or '')

    codon_list = make_codon_list(protein_seq, template_dna, include_stop)
    sanitize_codon_list(codon_list, forbidden_seqs)
    dna_seq = leading_seq + ''.join(codon_list) + trailing_seq

    if manufacturer == 'gen9':
        gen9.apply_quality_control_checks(dna_seq)

    return dna_seq

def make_codon_list(protein_seq, template_dna=None, include_stop=True):
    """
    Return a list of codons that would be translated to the given protein 
    sequence.  Codons are picked first to minimize the mutations relative to a 
    template DNA sequence and second to prefer "optimal" codons.
    """

    codon_list = []

    if template_dna is None:
        template_dna = []

    # Reverse translate each codon, preferring (in order):
    # 1. The codon with the most similarity to the template codon.
    # 2. The codon with the highest natural usage.

    for i, res in enumerate(protein_seq.upper()):
        try: template_codon = template_dna[3*i:3*i+3]
        except IndexError: template_codon = '---'

        # Already sorted by natural codon usage
        possible_codons = dna.ecoli_reverse_translate[res]

        # Sort by similarity.  Note that this is a stable sort.
        possible_codons.sort(
                key=lambda x: dna.num_mutations(x, template_codon))

        # Pick the best codon.
        codon_list.append(possible_codons[0])

    # Make sure the sequence ends with a stop codon.
    last_codon = codon_list[-1]
    stop_codons = dna.ecoli_reverse_translate['.']

    if include_stop and last_codon not in stop_codons:
        codon_list.append(stop_codons[0])

    return codon_list

def sanitize_codon_list(codon_list, forbidden_seqs=()):
    """
    Make silent mutations to the given codon lists to remove any undesirable 
    sequences that are present within it.  Undesirable sequences include 
    restriction sites, which may be optionally specified as a second argument, 
    and homopolymers above a pre-defined length.  The return value is the 
    number of corrections made to the codon list.
    """

    # Unit test missing for:
    #   Homopolymer fixing

    for codon in codon_list:
        if len(codon) != 3:
            raise ValueError("Codons must have exactly 3 bases: '{}'".format(codon))

    # Compile a collection of all the sequences we don't want to appear in the 
    # gene.  This includes the given restriction sites and their reverse 
    # complements, plus any homopolymers above a pre-defined length.

    bad_seqs = set()
    
    bad_seqs.union(
            restriction_sites.get(seq, seq)
            for seq in forbidden_seqs)

    bad_seqs.union(
            dna.reverse_complement(seq)
            for seq in bad_seqs)

    bad_seqs.union(
            base * (gen9.homopolymer_max_lengths[base] + 1)
            for base in dna.dna_bases)

    bad_seqs = [
            dna.dna_to_re(bs)
            for bs in bad_seqs]

    # Remove every bad sequence from the gene by making silent mutations to the 
    # codon list.
    
    num_corrections = 0

    for bad_seq in bad_seqs:
        while remove_bad_sequence(codon_list, bad_seq, bad_seqs):
            num_corrections += 1

    return num_corrections
        
def remove_bad_sequence(codon_list, bad_seq, bad_seqs):
    """
    Make a silent mutation to the given codon list to remove the first instance 
    of the given bad sequence found in the gene sequence.  If the bad sequence 
    isn't found, nothing happens and the function returns false.  Otherwise the 
    function returns true.  You can use these return values to easily write a 
    loop totally purges the bad sequence from the codon list.  Both the 
    specific bad sequence in question and the list of all bad sequences are 
    expected to be regular expressions.
    """

    gene_seq = ''.join(codon_list)
    problem = bad_seq.search(gene_seq)

    if not problem:
        return False

    bs_start_codon = problem.start() // 3
    bs_end_codon = problem.end() // 3

    for i in range(bs_start_codon, bs_end_codon):
        problem_codon = codon_list[i]
        amino_acid = translate_dna(problem_codon)

        alternate_codons = [
                codon
                for codon in dna.ecoli_reverse_translate[amino_acid]
                if codon != problem_codon]

        for alternate_codon in alternate_codons:
            codon_list[i] = alternate_codon

            if problem_with_codon(i, codon_list, bad_seqs):
                codon_list[i] = problem_codon
            else:
                return True

    raise RuntimeError("Could not remove bad sequence '{}' from gene.".format(bs))

def problem_with_codon(codon_index, codon_list, bad_seqs):
    """
    Return true if the given codon overlaps with a bad sequence.
    """

    base_1 = 3 * codon_index
    base_3 = 3 * codon_index + 2

    gene_seq = ''.join(codon_list)

    for bad_seq in bad_seqs:
        problem = bad_seq.search(gene_seq)
        if problem and problem.start() < base_3 and problem.end() > base_1:
            return True

    return False


def sequence_from_fasta(path):
    """
    Extract a single sequence from a FASTA file.

    It is an error for the FASTA file to contain more than one sequence.  If 
    you want to possibly load multiple sequences, use sequences_from_fasta() 
    instead.
    """
    from Bio import SeqIO
    return SeqIO.read(path, 'fasta').seq

def sequences_from_fasta(path):
    """
    Extract multiple sequences from a FASTA file.
    """
    from Bio import SeqIO
    return {x.description: x.seq for x in SeqIO.parse(path, 'fasta')}

def sequence_from_pdb(pdb, chain=None, ref=None):
    """
    Extract a protein sequence from a PDB file.
    
    Arguments
    =========
    pdb: str
        The path to a PDB file or a gzipped PDB file.  Alternatively, a PDB id 
        to fetch and download from the internet.

    chain: str
        Which chain or chains to extract the sequence from, and how to combine 
        sequences from multiple chains.  Different values for this option can 
        trigger quite different behaviors:

        - If no chain is specified (e.g. chain=None, the default), behavior 
          depends on how many chains are in the file.  If there's only one, its 
          sequence will be returned.  Otherwise a ValueError is raised and a 
          chain must be specified.

        - If a single chain is specified (e.g. chain="A"), the sequence of that 
          chain will be returned.

        - If multiple chains are joined by "+" signs (e.g. chain="A+B"), the 
          sequences for those chains will be concatenated together in the given 
          order.  This effectively combines the chains into one.
          
        - If multiple chains are joined by "~" signs (e.g. chain="A~B"), the 
          indicated chains will be considered as multiple copies of the same 
          protein, and mutations from every chain will be merged into one 
          sequence.  This behavior requires that the `ref` parameter is given.

        Note that mixing "+" and "~" is not currently allowed.

    ref: str
        The 

    Returns
    =======
    seq: str
        The extracted protein sequence (as a string of one-letter codes).
          
    Note that this function determines sequence from the ATOM records, not the 
    SEQRES records.  This means missing residues will not be included in the sequence.
    """

    seqs = chains_from_pdb(pdb)
    
    if chain is None:
        if len(seqs) == 1:
            return next(iter(seqs.values()))
        else:
            raise scripting.UserError("'{}' contains multiple chains, but no chain was specified.".format(pdb))

    single = re.match('^([A-Z])$', chain)
    merge = re.match('^(?:([A-Z])[~])+([A-Z])$', chain)
    concat = re.match('^(?:([A-Z])[+])+([A-Z])$', chain)
    
    if single:
        chain = single.group(1)
        return seqs[chain]

    if merge:
        if ref is None:
            raise ValueError('template required to merge sequences')

        mutants = [seqs[k] for k in merge.groups()]
        return merge_sequences(ref, *mutants)

    if concat:
        return ''.join(seqs[k] for k in concat.groups())

    raise ValueError("chain '{}' not understood".format(chain))

def chains_from_pdb(pdb):
    from Bio.PDB.PDBParser import PDBParser
    from Bio.PDB.Polypeptide import three_to_one, is_aa

    # Transparently handle gzipped PDB files.
    handle = pdb
    if str(pdb).endswith('.gz'):
        import io, gzip
        with gzip.open(pdb, 'rt') as file:
            handle = io.StringIO(file.read().decode('utf8'))

    parser = PDBParser(PERMISSIVE=1)
    structure = parser.get_structure(pdb, handle)

    seqs = {}
    for model in structure:
        for chain in model:
            seqs[chain.get_id()] = ''.join(
                    three_to_one(residue.get_resname())
                    for residue in chain
                    if is_aa(residue.get_resname())
            )

    return seqs

def merge_sequences(ref, *seqs):
    if not seqs:
        return ref

    if not all_seqs_same_length(seqs):
        raise ValueError("not all of the given sequences are the same length:\n" + '\n'.join(seqs))

    seq = ''

    for i in range(len(ref)):
        mutations = set(x[i] for x in seqs)
        mutations -= set([ref[i]])

        if len(mutations) == 0:
            seq += ref[i]
        elif len(mutations) == 1:
            seq += mutations.pop()
        else:
            raise ValueError("position {} has multiple mutations: {}".format(i+1, ','.join(mutations)))

    return seq

def all_seqs_same_length(seqs):
    return not seqs or all(len(x) == len(seqs[0]) for x in seqs[1:])


def write_sequences(path, seqs):
    """
    Write the given sequence to the given path, using the file format implied 
    by the path's file extension.

    Arguments
    =========
    path: str or pathlib.Path
        The name of the file to create.  Allowed extensions are:
        
        - *.fa, *.fas, *.fasta
        - *.tsv, *.csv
        - *.xlsx

    seqs: dict
        A mapping of names to sequences, which can be either protein or DNA.
    """
    path = Path(path)

    if path.suffix in ('.fa', '.fas', '.fasta'):
        write_sequences_to_fasta(path, seqs)

    elif path.suffix in ('.tsv', '.csv'):
        write_sequences_to_tsv(path, seqs)

    elif path.suffix in ('.xlsx',):
        write_sequences_to_xlsx(path, seqs)

    else:
        raise scripting.UserError("""\
'{}' has unknown output filetype '{}'.  Please provide a path with one of the 
following extensions:

- *.fa, *.fas, *.fasta
- *.tsv, *.csv
- *.xlsx
""".format(path, path.suffix))

def write_sequences_to_fasta(path, seqs):
    """
    Create a FASTA file listing the given sequences.

    Arguments
    =========
    path: str or pathlib.Path
        The name of the file to create.

    seqs: dict
        A mapping of names to sequences, which can be either protein or DNA.
    """
    from Bio import SeqIO
    from Bio.Seq import Seq
    from Bio.SeqRecord import SeqRecord

    path = Path(path)
    records = []

    for id, seq in seqs.items():
        record = SeqRecord(Seq(seq), id=id, description='')
        records.append(record)

    SeqIO.write(records, str(path), 'fasta')

def write_sequences_to_tsv(path, seqs):
    """
    Create a TSV (or CSV, depending on the extension) file listing the given 
    sequences.

    Arguments
    =========
    path: str or pathlib.Path
        The name of the file to create.  If the path extension is '.tsv', 
        fields will be delimited by tabs.  If the extension is '.csv', fields 
        will be delimited by commas.

    seqs: dict
        A mapping of names to sequences, which can be either protein or DNA.
    """
    import csv

    path = Path(path)
    delimiter = {'.tsv': '\t', '.csv': ','}[path.suffix]
    with path.open('w') as file:
        w = csv.writer(file, delimiter=delimiter)
        for row in seqs.items():
            w.writerow(row)
def write_sequences_to_xlsx(path, seqs):
    """
    Create a XLSX file listing the given sequences.

    Arguments
    =========
    path: str or pathlib.Path
        The name of the file to create.

    seqs: dict
        A mapping of names to sequences, which can be either protein or DNA.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    for row, id in enumerate(seqs, 1):
        ws.cell(row, 1).value = id
        ws.cell(row, 2).value = seqs[id]

    wb.save(path)
